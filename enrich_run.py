"""Refresh followers/engagement and recover bio emails for the outreach sheet."""
import csv, json, os, random, re, sys, time, urllib.request
sys.path.insert(0, '/home/user/Repo')
import openpyxl
from tiktok_scraper.enrich import extract_email

SRC = '/home/user/Repo/data/output/PrimalQueen_TikTok_Creators_REPLIES.xlsx'
OUT = '/home/user/Repo/data/output/enrichment_progress.csv'
DELAY = float(os.environ.get('REQUEST_DELAY_SECONDS', '2.5'))
UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

ws = openpyxl.load_workbook(SRC)['Creators (outreach)']
h = [c.value for c in ws[1]]
H, E, R, T = (h.index(x) for x in ('creator_handle', 'email', 'replied', 'tier'))

targets = []
for r in ws.iter_rows(min_row=2, values_only=True):
    handle = (r[H] or '').lstrip('@').strip()
    if not handle:
        continue
    targets.append({'handle': handle, 'old_email': (r[E] or '').strip(),
                    'replied': r[R], 'tier': r[T]})

# No email on file first -- that is where the value is.
targets.sort(key=lambda t: bool(t['old_email']))

done = set()
if os.path.exists(OUT):
    with open(OUT) as f:
        done = {row['handle'] for row in csv.DictReader(f)}

FIELDS = ['handle', 'status', 'followers', 'hearts', 'videos',
          'found_email', 'old_email', 'email_is_new', 'bio', 'tier', 'replied']
new = not os.path.exists(OUT)
fh = open(OUT, 'a', newline='')
w = csv.DictWriter(fh, fieldnames=FIELDS)
if new:
    w.writeheader()

def fetch(handle):
    req = urllib.request.Request(
        f'https://www.tiktok.com/@{handle}',
        headers={'User-Agent': UA, 'Accept-Language': 'en-US,en;q=0.9'})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode('utf-8', 'replace')

ok = fail = found = 0
for i, t in enumerate(targets, 1):
    if t['handle'] in done:
        continue
    rec = dict.fromkeys(FIELDS, '')
    rec.update(handle=t['handle'], old_email=t['old_email'],
               tier=t['tier'], replied=t['replied'], status='error')
    try:
        html = fetch(t['handle'])
        m = re.search(r'__UNIVERSAL_DATA_FOR_REHYDRATION__[^>]*>(.*?)</script>',
                      html, re.S)
        if not m:
            rec['status'] = 'blocked_or_no_data'
        else:
            info = (json.loads(m.group(1))['__DEFAULT_SCOPE__']
                    ['webapp.user-detail']['userInfo'])
            bio = info['user'].get('signature', '') or ''
            email = extract_email(bio)
            rec.update(status='ok',
                       followers=info['stats']['followerCount'],
                       hearts=info['stats']['heartCount'],
                       videos=info['stats']['videoCount'],
                       found_email=email, bio=bio.replace('\n', ' ')[:300],
                       email_is_new=bool(email and not t['old_email']))
            ok += 1
            if rec['email_is_new']:
                found += 1
    except Exception as exc:                       # noqa: BLE001
        rec['status'] = f'{type(exc).__name__}: {exc}'[:120]
        fail += 1
    w.writerow(rec)
    fh.flush()
    if i % 25 == 0:
        print(f'{i}/{len(targets)}  ok={ok} fail={fail} new_emails={found}',
              flush=True)
    time.sleep(DELAY + random.uniform(0, 1.0))

print(f'DONE  ok={ok} fail={fail} new_emails={found}', flush=True)
fh.close()
