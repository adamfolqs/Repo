"""Read the list of things to scrape out of a sheet you already have.

Accepts .csv / .xlsx and is deliberately forgiving about column naming --
your sheet probably calls it 'Creator', 'handle', 'TikTok URL' or similar.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

# Any of these column headers (case/space-insensitive) will be treated as
# the column holding the creator.
HANDLE_COLUMNS = {
    "handle", "handles", "username", "user", "creator", "creators",
    "account", "tiktok", "tiktokhandle", "tiktokurl", "profile", "profileurl", "url",
}

_HANDLE_RE = re.compile(r"tiktok\.com/@([A-Za-z0-9._]+)")


def normalize_handle(raw: str) -> str | None:
    """'@Nasa', 'https://tiktok.com/@nasa?x=1', 'nasa' -> 'nasa'."""
    if not raw:
        return None
    text = str(raw).strip()
    match = _HANDLE_RE.search(text)
    if match:
        return match.group(1)
    text = text.lstrip("@").strip()
    text = text.split("?")[0].split("/")[0]
    # A bare handle is letters/digits/._ only; anything else is a stray cell.
    return text if re.fullmatch(r"[A-Za-z0-9._]{1,24}", text) else None


def _key(name: str) -> str:
    return re.sub(r"[^a-z]", "", str(name).lower())


def read_handles(path: str | Path, column: str | None = None) -> list[str]:
    """Pull a deduped, ordered list of handles out of a csv/xlsx."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Input sheet not found: {path}")

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        rows = _read_xlsx(path)
    else:
        with path.open(newline="", encoding="utf-8-sig") as fh:
            rows = [r for r in csv.reader(fh)]

    if not rows:
        return []

    header, *body = rows
    target = None
    if column:
        wanted = _key(column)
        target = next((i for i, h in enumerate(header) if _key(h) == wanted), None)
        if target is None:
            raise ValueError(
                f"Column {column!r} not found. Available: {[h for h in header if h]}"
            )
    else:
        target = next((i for i, h in enumerate(header) if _key(h) in HANDLE_COLUMNS), None)

    handles: list[str] = []
    if target is None:
        # No recognizable header — scan every cell and keep what parses.
        # Covers the common case of a bare one-column list with no header.
        for row in rows:
            handles.extend(h for h in (normalize_handle(c) for c in row) if h)
    else:
        for row in body:
            if target < len(row):
                handle = normalize_handle(row[target])
                if handle:
                    handles.append(handle)

    seen, ordered = set(), []
    for handle in handles:
        low = handle.lower()
        if low not in seen:
            seen.add(low)
            ordered.append(handle)
    return ordered


def _read_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return [
        ["" if c is None else str(c) for c in row]
        for row in ws.iter_rows(values_only=True)
    ]
