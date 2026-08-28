"""Write rows to .csv and .xlsx on disk."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from pydantic import BaseModel


def _flatten(value: Any) -> Any:
    """Make a value safe for a spreadsheet cell."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, datetime):
        # Excel chokes on tz-aware datetimes; store a clean ISO string.
        return value.replace(tzinfo=None).isoformat(sep=" ", timespec="seconds")
    return value


def to_rows(records: Sequence[BaseModel], columns: Sequence[str]) -> list[list[Any]]:
    """Project models onto the fixed column order."""
    rows = []
    for rec in records:
        dumped = rec.model_dump()
        rows.append([_flatten(dumped.get(col)) for col in columns])
    return rows


def write_csv(records: Sequence[BaseModel], columns: Sequence[str], path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(columns)
        writer.writerows(to_rows(records, columns))
    return path


def write_xlsx(
    records: Sequence[BaseModel],
    columns: Sequence[str],
    path: str | Path,
    sheet_name: str = "Data",
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel's hard limit on tab names

    ws.append(list(columns))
    for row in to_rows(records, columns):
        ws.append(row)

    # Header styling + freeze, so the sheet is usable the moment it opens.
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Width by content, capped so a long caption doesn't blow out the sheet.
    for idx, col in enumerate(columns, start=1):
        longest = max(
            [len(str(col))] + [len(str(r[idx - 1])) for r in to_rows(records, columns)] or [0]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 10), 55)

    wb.save(path)
    return path
