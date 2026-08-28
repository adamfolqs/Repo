"""Recover creator contact emails from the links in their TikTok bios."""
import csv, gzip, io, os, random, re, socket, sys, time, urllib.error, urllib.request
sys.path.insert(0, '/home/user/Repo')
import openpyxl
from tiktok_scraper.enrich import extract_email

SRC = '/home/user/Repo/data/output/PrimalQueen_TikTok_Creators_REPLIES.xlsx'
OUT = '/home/user/Repo/data/output/contacts_found.csv'
DELAY = float(os.environ.get('REQUEST_DELAY_SECONDS', '2.0'))
# urllib's per-call timeout does not cover DNS or a stalled TLS handshake, so
# a dead creator domain can hang the whole run. This bounds every socket op.
socket.setdefaulttimeout(12)

UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# Aggregators worth following one hop into; skip storefronts and dead ends.
SKIP = re.compile(r'(amazon\.|paypal\.me|primalqueen\.com|tiktok\.com/|'
                  r'youtube\.com|venmo\.com|cash\.app)', re.I)
# Junk addresses that belong to the platform, not the creator.
JUNK = re.compile(r'(sentry|wixpress|example\.|@linktr|@beacons|@stan\.|'
                  r'@msha|@tiktok|@sentry\.io|\.png|\.jpg|\.webp|@2x)', re.I)

ws = openpyxl.load_workbook(SRC)['Creators (outreach)']
h = [c.value for c in ws[1]]
H, E, L, F, T = (h.index(x) for x in
                 ('creator_handle', 'email', 'bio_link', 'followers', 'tier'))

targets = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if (r[E] or '').strip():
        continue
    link = (r[L] or '').strip()
    if not link or SKIP.search(link):
        continue
    targets.append({'handle': r[H], 'link': link,
                    'followers': r[F], 'tier': r[T]})
targets.sort(key=lambda t: -(int(t['followers'] or 0)))

done = set()
if os.path.exists(OUT):
    with open(OUT) as f:
        done = {row['handle'] for row in csv.DictReader(f)}

FIELDS = ['handle', 'followers', 'tier', 'link', 'status', 'found_email', 'source']
new = not os.path.exists(OUT)
fh = open(OUT, 'a', newline='')
w = csv.DictWriter(fh, fieldnames=FIELDS)
if new:
    w.writeheader()

def get(url, timeout=12):
    req = urllib.request.Request(url, headers={
        'User-Agent': UA, 'Accept-Language': 'en-US,en;q=0.9',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
        'Accept-Encoding': 'gzip'})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    if raw[:2] == b'\x1f\x8b':
        raw = gzip.GzipFile(fileobj=io.BytesIO(raw)).read()
    return raw.decode('utf-8', 'replace')

MAILTO = re.compile(r'mailto:([^"\'?&<>\s]+)', re.I)
# The domain is matched as explicit labels rather than [A-Za-z0-9.-]+\. --
# that form is ambiguous with the following dot and backtracks catastrophically
# on big minified pages (a 1.3 MB page hung for minutes). Bounded and anchored
# on a word boundary, this stays linear.
PLAIN = re.compile(r'[A-Za-z0-9._%+-]{1,64}@[A-Za-z0-9-]{1,63}'
                   r'(?:\.[A-Za-z0-9-]{1,63}){1,3}\b')
# Guard against pathological pages regardless; contact details live near the
# top of a link-in-bio page, never 500 KB deep.
MAX_SCAN = 500_000

def emails_from(html):
    html = html[:MAX_SCAN]
    hits = []
    for m in MAILTO.findall(html):
        hits.append(m)
    for m in PLAIN.findall(html):
        hits.append(m)
    out = []
    for e in hits:
        e = e.strip('.,;:"\'').lower()
        if JUNK.search(e) or len(e) > 80:
            continue
        out.append(e)
    return out

found = 0
for i, t in enumerate(targets, 1):
    if t['handle'] in done:
        continue
    rec = {'handle': t['handle'], 'followers': t['followers'], 'tier': t['tier'],
           'link': t['link'], 'status': '', 'found_email': '', 'source': ''}
    try:
        url = t['link'] if t['link'].startswith('http') else 'https://' + t['link']
        html = get(url)
        hits = emails_from(html)
        if not hits:
            # Linktree etc. render links client-side; the JSON blob still
            # carries any mailto the creator added.
            hits = [e.strip('.,;:"\'').lower()
                    for e in re.findall(r'mailto:\\?u?0?0?3?a?([^"\\\s<>]+)',
                                        html[:MAX_SCAN])
                    if not JUNK.search(e)]
        if hits:
            rec.update(status='ok', found_email=hits[0], source=url)
            found += 1
        else:
            rec['status'] = 'no_email_on_page'
    except urllib.error.HTTPError as exc:
        rec['status'] = f'HTTP {exc.code}'
    except Exception as exc:                        # noqa: BLE001
        rec['status'] = f'{type(exc).__name__}'[:40]
    w.writerow(rec)
    fh.flush()
    print(f"{i}/{len(targets)} {rec['handle']} {rec['status']} "
          f"{rec['found_email']}", flush=True)
    time.sleep(DELAY + random.uniform(0, 0.8))

print(f'DONE  checked={len(targets)} emails_found={found}', flush=True)
fh.close()
