"""Score recovered emails and merge the usable ones into the outreach sheet."""
import csv, re
import openpyxl
from openpyxl.styles import PatternFill

SHEET = '/home/user/Repo/data/output/PrimalQueen_TikTok_Creators_REPLIES.xlsx'
FOUND = '/home/user/Repo/data/output/contacts_found.csv'

# Placeholders and third-party brands the scraper picked up off the page
# rather than from the creator's own contact block.
REJECT = re.compile(r'(@domain\.com|@example\.|@\d+\.\d+|@comfrt\.com|'
                    r'noreply|no-reply|@sentry|@wix)', re.I)
FREEMAIL = ('gmail.com', 'yahoo.com', 'outlook.com', 'aol.com',
            'icloud.com', 'hotmail.com', 'proton.me')

def norm(s):
    return re.sub(r'[^a-z0-9]', '', (s or '').lower())

def score(handle, name, email):
    """HIGH when the address clearly belongs to this creator."""
    local, _, domain = email.partition('@')
    h, n, l = norm(handle), norm(name), norm(local)
    stem = norm(domain.split('.')[0])
    # Address echoes the handle or display name -> almost certainly theirs.
    if len(l) >= 4 and (l in h or h in l or (n and (l in n or n in l))):
        return 'HIGH'
    if domain not in FREEMAIL and len(stem) >= 4 and (stem in h or h in stem
                                                      or (n and stem in n)):
        return 'HIGH'
    if domain in FREEMAIL:
        return 'MEDIUM'
    return 'LOW'          # generic address on an unrelated domain

rows = [r for r in csv.DictReader(open(FOUND)) if r['found_email']]
wb = openpyxl.load_workbook(SHEET)
ws = wb['Creators (outreach)']
h = [c.value for c in ws[1]]
H, E, N = h.index('creator_handle'), h.index('email'), h.index('creator_name')
if 'email_confidence' not in h:
    ws.cell(row=1, column=len(h) + 1, value='email_confidence')
    ws.cell(row=1, column=len(h) + 2, value='email_source')
CONF, SRC = len(h), len(h) + 1

by_handle = {}
for r in rows:
    if REJECT.search(r['found_email']):
        r['verdict'] = 'REJECT'
    else:
        r['verdict'] = score(r['handle'], '', r['found_email'])
    by_handle[r['handle']] = r

fill = PatternFill('solid', fgColor='DDEBF7')
merged = 0
for row in ws.iter_rows(min_row=2):
    r = by_handle.get(row[H].value)
    if not r or r['verdict'] == 'REJECT' or (row[E].value or '').strip():
        continue
    # Re-score now that the display name is available.
    r['verdict'] = score(r['handle'], row[N].value, r['found_email'])
    row[E].value = r['found_email'].rstrip('\\/.,;')
    row[CONF].value = r['verdict']
    row[SRC].value = r['link']
    for c in row[:len(h) + 2]:
        c.fill = fill
    merged += 1
wb.save(SHEET)

from collections import Counter
print('merged into sheet :', merged)
print('verdicts          :', dict(Counter(r['verdict'] for r in by_handle.values())))
print()
for r in sorted(by_handle.values(), key=lambda x: -int(x['followers'] or 0)):
    print(f"  {r['verdict']:7} {r['handle'][:28]:30} {int(r['followers'] or 0):>9,}  {r['found_email']}")
