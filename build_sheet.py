"""Assemble the deliverable workbook from the scrape stores.

Reads the append-only JSONL stores written by the discover/resolve/catalogue/
profiles stages and produces `data/output/colostrum_creator_list.xlsx`.

Two rules run through the whole file:

* No number is ever invented. Every like/view/follower count came from the
  record TikTok served for that exact video or profile. A field we could not
  read stays empty, because a plausible-looking estimate would silently become
  the sort order for outreach decisions.
* No handle is ever guessed from a display name. A row in the original
  sourcing sheet only gets a handle when a creator we actually landed on
  matches it *and* that creator is in the colostrum dataset.
"""

from __future__ import annotations

import csv
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tiktok_scraper import store
from tiktok_scraper.enrich import enrich_creators, enrich_videos
from tiktok_scraper.models import Creator, Video

OUT = Path("data/output/colostrum_creator_list.xlsx")
ORIGINAL = Path("data/input/colostrum_sourcing.xlsx")

MIN_LIKES = 50


# --------------------------------------------------------------------- load

def load_videos() -> list[Video]:
    """Every resolved video, newest record per id winning."""
    by_id: dict[str, Video] = {}
    for record in store.read("data/output/videos.jsonl"):
        try:
            video = Video(**record)
        except Exception:
            continue
        by_id[video.video_id] = video
    return list(by_id.values())


def load_creators() -> list[Creator]:
    by_handle: dict[str, Creator] = {}
    for record in store.read("data/output/creators.jsonl"):
        try:
            creator = Creator(**record)
        except Exception:
            continue
        by_handle[creator.handle.lower()] = creator
    return list(by_handle.values())


# ------------------------------------------------------------------ styling

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style(ws, widths: dict[str, int] | None = None, wrap_cols: tuple[str, ...] = ()) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
        )
    for column in ws.columns:
        letter = column[0].column_letter
        header = str(ws.cell(row=1, column=column[0].column).value or "")
        if widths and header in widths:
            ws.column_dimensions[letter].width = widths[header]
            continue
        longest = max((len(str(c.value or "")) for c in column[:200]), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 46)
    for header in wrap_cols:
        for column in ws.columns:
            if str(ws.cell(row=1, column=column[0].column).value or "") == header:
                for cell in column[1:]:
                    cell.alignment = Alignment(wrap_text=False, vertical="top")


def write_tab(wb: Workbook, title: str, columns: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append(row)
    style(ws)


def video_row(video: Video, columns: list[str]) -> list:
    out = []
    for column in columns:
        value = getattr(video, column, None)
        if isinstance(value, list):
            value = ", ".join(str(v) for v in value)
        elif hasattr(value, "isoformat"):
            value = value.replace(tzinfo=None).isoformat(sep=" ", timespec="minutes")
        out.append(value)
    return out


# --------------------------------------------------- original-sheet matching

def normalize_name(value: str | None) -> str:
    """Loose key for comparing a display name to a handle or nickname."""
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return "".join(c for c in text.lower() if c.isalnum())


def match_original_sheet(videos, creators):
    """Fill handles into the original sourcing sheet, evidence-only.

    A name from the screen recording is matched against creators we actually
    scraped -- by handle or by display name -- and accepted only when that
    creator is in the colostrum dataset. Generic first names ('Taylor',
    'Nikki') are left unresolved on purpose: several real accounts share them,
    so any pick would be a guess, and a wrong handle sends outreach to a
    stranger.
    """
    by_key: dict[str, list[Creator]] = defaultdict(list)
    for creator in creators:
        for key in {normalize_name(creator.handle), normalize_name(creator.nickname)}:
            if key:
                by_key[key].append(creator)

    brands_by_handle: dict[str, set[str]] = defaultdict(set)
    for video in videos:
        if video.competitor_brand:
            brands_by_handle[video.handle.lower()].add(video.competitor_brand)
    colostrum_handles = {
        v.handle.lower() for v in videos if v.is_colostrum
    }

    # Handles confirmed by the previous session's manual lookup.
    prior: dict[str, dict] = {}
    lookup_path = Path("data/output/handle_lookup.csv")
    if lookup_path.exists():
        with lookup_path.open(encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                if row.get("handle"):
                    prior[normalize_name(row.get("name"))] = row

    # Handles verified this run by opening the profile (see handles.py).
    verified: dict[str, dict] = {}
    resolution_path = Path("data/output/name_resolution.csv")
    if resolution_path.exists():
        with resolution_path.open(encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                verified[normalize_name(row.get("name"))] = row

    wb = load_workbook(ORIGINAL, data_only=True)
    ws = wb["Creator Videos"]
    rows = list(ws.iter_rows(values_only=True))
    header, body = rows[0], [r for r in rows[1:] if any(r)]
    index = {name: i for i, name in enumerate(header)}

    out, filled = [], 0
    for row in body:
        name = row[index["Creator label from recording"]]
        brand = row[index["Product / Brand"]]
        key = normalize_name(name)

        handle = profile = status = evidence = ""

        hit = prior.get(key)
        if hit:
            handle = hit["handle"]
            profile = hit.get("profile_url") or f"https://www.tiktok.com/@{handle}"
            status = hit.get("confidence") or "confirmed"
            evidence = hit.get("evidence") or "manual lookup, previous session"
            # Two accounts can carry the same display name. Where this run
            # landed on a different one, say so rather than quietly keeping
            # either -- the earlier lookup has video evidence behind it, so it
            # wins, but the disagreement is worth a human glance.
            other = verified.get(key, {}).get("handle")
            if other and other.lower() != handle.lower():
                status = "confirmed, but a second account shares this name"
                evidence += f" | also verified: @{other} displays the same name"

        if not handle and key in verified:
            row_hit = verified[key]
            if row_hit.get("handle"):
                handle = row_hit["handle"]
                profile = row_hit.get("profile_url") or f"https://www.tiktok.com/@{handle}"
                status = row_hit.get("confidence") or "confirmed"
                evidence = row_hit.get("evidence") or ""
            elif row_hit.get("confidence"):
                status = row_hit["confidence"]
                evidence = row_hit.get("evidence") or ""

        if not handle:
            candidates = [
                c for c in by_key.get(key, [])
                if c.handle.lower() in colostrum_handles
            ]
            if len(candidates) == 1:
                creator = candidates[0]
                handle = creator.handle
                profile = creator.profile_url
                shared = brand and brand in brands_by_handle.get(handle.lower(), set())
                status = "confirmed (scraped)"
                evidence = (
                    f"name matches @{handle}"
                    + (f"; posts about {brand}" if shared else "")
                    + f"; {sum(1 for v in videos if v.handle.lower() == handle.lower())}"
                    " colostrum video(s) in this scrape"
                )
            elif len(candidates) > 1:
                status = "ambiguous — several accounts match this name"
                evidence = "candidates: " + ", ".join(f"@{c.handle}" for c in candidates[:5])
            else:
                status = "not found"

        if handle:
            filled += 1
        out.append([
            brand, name, handle, profile, status,
            row[index["Likes"]], row[index["Engagement tier"]],
            row[index["OCR confidence"]], evidence,
        ])
    return out, filled


ORIGINAL_HEADER = [
    "Brand", "Name from recording", "Handle", "Profile URL", "Lookup status",
    "Likes (from recording)", "Engagement tier", "OCR confidence", "Evidence",
]


# ---------------------------------------------------------------------- main

def main() -> int:
    videos = load_videos()
    creators = load_creators()
    if not videos:
        print("No videos in the store yet — run discover/resolve first.", file=sys.stderr)
        return 1

    enrich_creators(creators)
    enrich_videos(videos, creators)

    colostrum = [v for v in videos if v.is_colostrum]
    qualifying = [
        v for v in colostrum
        if (v.likes or 0) >= MIN_LIKES and v.has_product_tag
    ]
    qualifying.sort(key=lambda v: (v.likes or 0), reverse=True)
    skeptical = sorted(
        (v for v in colostrum if v.stance == "skeptical"),
        key=lambda v: (v.likes or 0), reverse=True,
    )
    colostrum_sorted = sorted(colostrum, key=lambda v: (v.likes or 0), reverse=True)

    video_columns = [
        "video_url", "handle", "creator_followers", "creator_email",
        "likes", "views", "comments", "shares", "saves", "engagement_rate",
        "language", "competitor_brand", "stance", "brand_account",
        "is_colostrum", "has_product_tag",
        "description", "hashtags", "music_title", "created_at",
        "duration_seconds", "matched_query", "video_id", "source",
    ]

    # ---- creators, aggregated over their videos --------------------------
    by_handle = {c.handle.lower(): c for c in creators}
    agg: dict[str, dict] = defaultdict(
        lambda: {"videos": 0, "qualifying": 0, "brands": set(), "langs": set(),
                 "likes": 0, "skeptical": 0, "best": None}
    )
    for video in colostrum:
        entry = agg[video.handle.lower()]
        entry["videos"] += 1
        entry["likes"] += video.likes or 0
        if video.competitor_brand:
            entry["brands"].add(video.competitor_brand)
        if video.language and video.language != "unknown":
            entry["langs"].add(video.language)
        if video.stance == "skeptical":
            entry["skeptical"] += 1
        if (video.likes or 0) >= MIN_LIKES and video.has_product_tag:
            entry["qualifying"] += 1
        if not entry["best"] or (video.likes or 0) > (entry["best"].likes or 0):
            entry["best"] = video

    creator_columns = [
        "handle", "profile_url", "followers", "email", "language",
        "brand_account", "colostrum_videos", "qualifying_videos",
        "total_likes_on_colostrum_videos", "brands_featured", "skeptical_videos",
        "top_video_url", "top_video_likes",
        "nickname", "verified", "total_likes", "video_count", "region",
        "bio", "bio_link",
    ]
    creator_rows = []
    for handle, entry in sorted(
        agg.items(), key=lambda kv: (kv[1]["qualifying"], kv[1]["likes"]), reverse=True
    ):
        creator = by_handle.get(handle)
        best = entry["best"]
        creator_rows.append([
            handle,
            f"https://www.tiktok.com/@{handle}",
            creator.followers if creator else None,
            (creator.email if creator else None) or "",
            (creator.language if creator and creator.language != "unknown" else None)
            or ", ".join(sorted(entry["langs"])) or "unknown",
            "YES" if (creator and creator.brand_account) else "",
            entry["videos"], entry["qualifying"], entry["likes"],
            ", ".join(sorted(entry["brands"])),
            entry["skeptical"] or "",
            best.video_url if best else "",
            best.likes if best else None,
            creator.nickname if creator else None,
            "YES" if (creator and creator.verified) else "",
            creator.total_likes if creator else None,
            creator.video_count if creator else None,
            creator.region if creator else None,
            (creator.bio if creator else None) or "",
            (creator.bio_link if creator else None) or "",
        ])

    original_rows, filled = match_original_sheet(colostrum, creators)

    # ---- write -----------------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)
    write_tab(wb, "Videos", video_columns,
              [video_row(v, video_columns) for v in qualifying])
    write_tab(wb, "Creators", creator_columns, creator_rows)
    write_tab(wb, "Original Sheet + Handles", ORIGINAL_HEADER, original_rows)
    write_tab(wb, "Objection Research", video_columns,
              [video_row(v, video_columns) for v in skeptical])
    write_tab(wb, "All Colostrum Videos", video_columns,
              [video_row(v, video_columns) for v in colostrum_sorted])
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    # ---- report ----------------------------------------------------------
    langs: dict[str, int] = defaultdict(int)
    brands: dict[str, int] = defaultdict(int)
    for video in qualifying:
        langs[video.language or "unknown"] += 1
        brands[video.competitor_brand or "(none named)"] += 1
    with_email = sum(1 for row in creator_rows if row[3])
    with_followers = sum(1 for row in creator_rows if row[2])

    print(f"wrote {OUT}")
    print(f"  Videos (colostrum, {MIN_LIKES}+ likes, product-tagged): {len(qualifying)}")
    print(f"  Creators                                  : {len(creator_rows)}")
    print(f"  Original Sheet + Handles                  : {len(original_rows)} "
          f"({filled} with a handle)")
    print(f"  Objection Research                        : {len(skeptical)}")
    print(f"  All Colostrum Videos                      : {len(colostrum_sorted)}")
    print(f"\n  resolved videos in store: {len(videos)}  profiles: {len(creators)}")
    print(f"  creators with follower count: {with_followers}, with email: {with_email}")
    print(f"\n  language split (Videos tab): {dict(langs)}")
    print("  brand split (Videos tab):")
    for brand, count in sorted(brands.items(), key=lambda kv: -kv[1]):
        print(f"    {brand:22} {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
